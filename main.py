import wx
import wx.grid
from openpyxl import Workbook,load_workbook
from openpyxl.styles import *
import warnings
warnings.filterwarnings('ignore')


from openpyxl import load_workbook
from openpyxl.styles import Border,Side
wb = load_workbook("模板.xlsx")#使用openpyxl读取xlsx文件，创建workbook
ws = wb.active

from openpyxl.styles import Border, Side
border_type=Side(border_style=None, color='FF000000')
border = Border(left=border_type,
right=border_type,
top=border_type,
bottom=border_type,
diagonal=border_type,
diagonal_direction=0,
outline=border_type,
vertical=border_type,
horizontal=border_type)



class GridFrame(wx.Frame):
    def __init__(self, parent):
        wx.Frame.__init__(self, parent)

        grid = wx.grid.Grid(self, -1)

        grid.CreateGrid(50, 10)

        # grid.SetRowSize(0, 60)
        # grid.SetColSize(0, 60)

        grid.SetCellValue(0, 0, '姓名')
        grid.SetCellTextColour(0, 0, wx.BLACK)
        grid.SetCellBackgroundColour(0, 0, wx.LIGHT_GREY)

        grid.SetCellValue(0, 1, '班级')
        grid.SetCellTextColour(0, 1, wx.YELLOW)
        grid.SetCellBackgroundColour(0, 1, wx.LIGHT_GREY)
        grid.SetReadOnly(0, 1)

        grid.SetCellValue(0,2, '学号')
        grid.SetCellTextColour(0, 2, wx.BLUE)
        grid.SetCellBackgroundColour(0, 2, wx.LIGHT_GREY)

        grid.SetColFormatFloat(5, 6, 2)
        grid.SetCellValue(1, 0, '郑俊杰')
        grid.SetCellTextColour(1, 0, wx.RED)
        grid.SetCellValue(1, 1, '物联网二班')
        grid.SetCellTextColour(1, 2, wx.WHITE)
        grid.SetCellValue(1, 2, '202011909')
        grid.SetCellTextColour(1, 2, wx.YELLOW)



        self.Show()

if __name__ == '__main__':

    app = wx.App(0)
    frame = GridFrame(None)
    app.MainLoop()
